import pandas as pd
import openpyxl as pyxl
import os

class excel_extract(object):
    def __init__(self,file_path,sheet_name=None):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def diy_sheet(self,sheet_name):
        if not sheet_name:
            return self.sheet_name
        else:
            return sheet_name

    def data_cleaner(data):
        data_no_na = data.dropna()
        data_cleaned = data_no_na.astype(str).str.replace(' ', '')
        return data_cleaned
    
    def col_byname(self,column_name,sheet_name=None):
        if not sheet_name:
            sheet_name=self.sheet_name
        # 使用pandas读取excel文件
        df = pd.read_excel(self.file_path,sheet_name)
        # 提取特定列，这里通过列名来索引
        column = df[column_name]
        #去除空值和空格
        specific_column_cleaned = excel_extract.data_cleaner(column)
        # 将提取的列转换为列表
        column_as_list = [specific_column_cleaned.tolist(),column.tolist()]
        
        return column_as_list   

    def row(self,row_number,sheet_name=None):
        if not sheet_name:
            sheet_name=self.sheet_name
        # 使用pandas读取excel文件
        df = pd.read_excel(self.file_path,sheet_name)
        # 提取特定行，这里使用iloc，因为我们是按照行号来索引
        # 注意：行号是从0开始的，所以如果你想要提取第5行，row_number应该是4
        specific_row = df.iloc[row_number-1]
        cleaned_row = excel_extract.data_cleaner(specific_row)
        # 将提取的行转换为列表
        row_as_list = [cleaned_row.tolist(),specific_row.tolist()]
        
        return row_as_list
    
    def col_byindex(self,column_index,sheet_name=None):
    # 使用pandas读取excel文件，指定工作表名称或索引
        if not sheet_name:
            sheet_name=self.sheet_name
        df = pd.read_excel(self.file_path, sheet_name)
        # 根据列的索引提取特定列
        # 注意：这里假设DataFrame的列足够多，以包含该索引
        specific_column = df.iloc[:, column_index-1]
        # 将非字符串值转换为字符串，并去除空格
        specific_column_cleaned = excel_extract.data_cleaner(specific_column)
        # 将提取的列转换为列表
        column_as_list = [specific_column_cleaned.tolist(),specific_column.tolist()]
        
        return column_as_list

class excel_writer(object):
    def __init__(self,file_path,sheet_name=None):
        self.file_path=file_path
        self.sheet_name=sheet_name
    
    def diy_sheet(self,sheet_name):
        if not sheet_name:
            return self.sheet_name
        else:
            return sheet_name
            
    def diy_excel_path(self,excel_path):
        if not excel_path:
            return self.file_path
        else:
            return excel_path
                
    def write_list_to_row(self, row_number, data_list,excel_path=None, sheet_name=None):
        sheet_name=self.diy_sheet(sheet_name)
        excel_path=self.diy_excel_path(excel_path)
        
        if not os.path.exists(excel_path):
            wb = pyxl.Workbook()
            wb.create_sheet(sheet_name)
            wb.save(excel_path)
            wb.close()
        
        workbook = pyxl.load_workbook(filename=excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]
        
        for col, value in enumerate(data_list, start=1):
            sheet.cell(row=row_number, column=col, value=value)
        
        workbook.save(excel_path)
        workbook.close()

    def write_list_to_column(self, column_number, data_list,excel_path=None, sheet_name=None):
        # 如果未指定工作表名称，则使用初始化时指定的名称
        sheet_name=self.diy_sheet(sheet_name)
        excel_path=self.diy_excel_path(excel_path)
        
        # 加载或创建工作簿和工作表
        workbook = pyxl.load_workbook(excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]
        
        # 将数据列表写入指定的列
        for row, value in enumerate(data_list, start=2):
            if not isinstance(value, (str, bytes)):
                value = str(value)
            sheet.cell(row=row, column=column_number, value=value)
        
        # 保存并关闭工作簿
        workbook.save(excel_path)
        workbook.close()

if __name__=='__main__':
    file_path=r"D:\python code\sort_docx\files\分类结果.xlsx"
    sheet_name='Sheet1'
    excel=excel_extract(file_path,sheet_name)
    excel1=excel_writer(file_path,sheet_name)
    list1=excel.col_byindex(1,None)[0][0:10]
    
    #print(excel.col_byname('姓名','总表')[0])
    #print(excel.row(2,'总表')[0])
    #excel1.write_list_to_row(1,excel.col_byindex(1,None)[0],r"C:\Users\Haryevsky\Desktop\新建 XLSX 工作表1.xlsx")
    #excel1.write_list_to_column(r"C:\Users\11\Desktop\test.xlsx",1,excel.col_byname('姓名','总表')[1])